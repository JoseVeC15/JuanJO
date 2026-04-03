import openpyxl
f = r"D:\Tam\TAMM TAMM\ALEJANDRA CASTEL\IVA 2025\01_Enero 2025\AC_Liquidación de IVA_012025.xlsx"
try:
    wb = openpyxl.load_workbook(f, data_only=False)
    for sn in wb.sheetnames:
        print(f"\n- {sn}")
        s = wb[sn]
        for r in s.iter_rows(max_row=30, max_col=10):
            row = [f"{c.coordinate}:{c.value}" for c in r if c.value]
            if row: print(" | ".join(row))
except Exception as e: print(e)
